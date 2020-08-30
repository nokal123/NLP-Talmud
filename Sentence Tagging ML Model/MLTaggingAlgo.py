# File: Tagging Algo

import nltk
import numpy
import openpyxl
import pickle
from nltk.tokenize import wordpunct_tokenize
from nltk import bigrams, trigrams

uni, bi, tri = pickle.load(open("bag_of_words.p", "rb"))
all_words = uni
word_features = list(all_words)


def gemara_features(words):
    features = {}
    punc = (words[-1])[-1]
    features['punctuation'] = punc
    the_bigrams = list(bigrams(words))
    the_trigrams = list(trigrams(words))

    for word in word_features:
        features['contains({})'.format(word)] = (word in words)

    for b in bi:
        features['contains({})'.format(b)] = (b in the_bigrams)

    for t in tri:
        features['contains({})'.format(t)] = (t in the_trigrams) # was format(tri)

    return features


# Goal will be to use the Tsv file instead of the hand-made xlsx file, Berakhot.xlsx.
# Current issues on edge cases have prevented its use. Code that will in the future be used is:
# wb = openpyxl.load_workbook('TsvBerakhot.xlsx')

wb = openpyxl.load_workbook('T.xlsx')
sheet = wb["daf_1"]

# Load in dictionary from Pickle file
diction = pickle.load(open("wordDict.p", "rb"))
prev_cell = ''

labeled_sentences = []
words = []
# reading in tagged data set used to train model
for row_num in range(sheet.max_row):
    phrase = sheet.cell(row_num + 1, 1).value
    if sheet.cell(row_num + 1, 2).value and sheet.cell(row_num + 2, 2).value is None:
        break
    words = wordpunct_tokenize(phrase) #this belongs becuase of features
    labeled_sentences.append((words, sheet.cell(row_num + 1, 2).value))


featureset = [(gemara_features(phrase), label) for (phrase,label) in labeled_sentences]
SPLIT_POS = 50
train_set, test_set = featureset[SPLIT_POS:], featureset[:SPLIT_POS]
train_sentences, test_sentences = labeled_sentences[SPLIT_POS:], labeled_sentences[:SPLIT_POS]

# Different classifiers where tested for this model. Others used where:
#   classifier = nltk.classify.NaiveBayesClassifier.train(train_set)
#   classifier = nltk.classify.MaxentClassifier.train(train_set)
classifier = nltk.classify.MaxentClassifier.train(train_set, algorithm='gis', max_iter=5)
print(nltk.classify.accuracy(classifier, test_set))
errors = []
for (sentence_features, tag), (sentence, _) in zip(test_set, test_sentences):
    guess = classifier.classify(sentence_features)
    if guess != tag:
        errors.append((tag, guess, sentence))

for (tag, guess, sentence) in sorted(errors):
    print('correct={:<8} guess={:<8s} sentence={}'.format(tag, guess, sentence))

classifier.show_most_informative_features(10)


wb.save('Tagged_Gemara.xlsx')
wb.close()
